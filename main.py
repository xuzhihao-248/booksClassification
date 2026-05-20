#!/usr/bin/env python3
"""
中文图书数据采集系统 — 多源版 (豆瓣读书 + 当当网)
输出: data/output/books_100k.xlsx  (书名 / 内容简介 / 中图法分类号 三栏核心数据)

用法:
    python main.py                        # 正常运行 (目标 100000 条)
    python main.py --max 100              # 测试运行 (100 条)
    python main.py --max 5000             # 中等运行 (5000 条)
    python main.py --resume               # 从断点恢复
"""

import argparse
import json
import os
import random
import re
import sqlite3
import sys
import time
import logging
from collections import Counter
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook

# ── 路径配置 ──────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = DATA_DIR / "output"
LOG_DIR = BASE_DIR / "logs"
CHECKPOINT_FILE = DATA_DIR / "checkpoint.json"
SEEN_DB = DATA_DIR / "seen_books.db"
TAG_MAP_FILE = BASE_DIR / "tag_to_clc.json"
OUTPUT_FILE = OUTPUT_DIR / "books_100k.xlsx"

for d in [DATA_DIR, OUTPUT_DIR, LOG_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ── 日志 ──────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "scraper.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("book_scraper")

# ── 常量 ──────────────────────────────────────────────────
DOUBAN_BASE = "https://book.douban.com"
DOUBAN_TAG = f"{DOUBAN_BASE}/tag"
DOUBAN_SUBJECT = f"{DOUBAN_BASE}/subject"

DANGDANG_BASE = "http://category.dangdang.com"
DANGDANG_PRODUCT = "http://product.dangdang.com"
DANGDANG_SEARCH = "http://search.dangdang.com"

REQUEST_DELAY_MIN = 2.0
REQUEST_DELAY_MAX = 5.0
RETRY_BASE_DELAY = 10.0
MAX_RETRIES = 5
MAX_BOOKS_DEFAULT = 100_000
BATCH_SIZE = 500
CHECKPOINT_INTERVAL = 1000

# ── 中图法目标分布 ────────────────────────────────────────
CATEGORY_TARGETS = {
    "A": 1500, "B": 5000, "C": 5000, "D": 4000, "E": 2500,
    "F": 5000, "G": 4000, "H": 3500, "I": 8000, "J": 5000,
    "K": 6000, "N": 3000, "O": 3000, "P": 2000, "Q": 3000,
    "R": 4000, "S": 2500, "T": 7000, "U": 1500, "V": 1500,
    "X": 2000, "Z": 4000,
}

# ── CLC类别 → 豆瓣标签 (用于URL发现) ──────────────────────
CATEGORY_TAGS = {
    "A": ["马克思主义", "毛泽东思想", "邓小平理论", "资本论"],
    "B": ["哲学", "心理学", "宗教", "佛教", "伦理学", "美学", "逻辑学", "西方哲学", "中国哲学", "道德经"],
    "C": ["社会学", "人类学", "统计学", "管理学", "人口学", "民族学", "社会心理学"],
    "D": ["政治", "法律", "国际关系", "政治学", "法学", "宪法", "刑法", "民法"],
    "E": ["军事", "战争", "二战", "战略", "兵器", "孙子兵法", "军事史"],
    "F": ["经济学", "金融", "投资", "市场营销", "会计", "创业", "股票", "理财", "贸易", "商业"],
    "G": ["教育", "文化", "传播", "新闻", "体育", "教育学", "传播学", "出版", "足球", "篮球"],
    "H": ["英语", "日语", "语言学", "翻译", "法语", "德语", "词典", "语法", "汉字"],
    "I": ["小说", "诗歌", "散文", "戏剧", "外国文学", "中国文学", "名著", "武侠", "科幻", "推理", "童话", "儿童文学", "日本文学", "英国文学", "美国文学", "法国文学", "俄罗斯文学"],
    "J": ["艺术", "绘画", "摄影", "电影", "音乐", "设计", "书法", "建筑", "漫画", "古典音乐", "艺术史", "电影理论"],
    "K": ["历史", "中国历史", "世界史", "传记", "地理", "考古", "欧洲史", "日本史", "旅行", "游记", "二战史"],
    "N": ["科普", "自然科学", "科学史", "科学哲学", "系统科学"],
    "O": ["数学", "物理", "化学", "量子力学", "微积分", "代数", "几何", "概率论", "相对论"],
    "P": ["天文学", "地质", "气象", "海洋", "宇宙", "地球科学", "古生物", "黑洞"],
    "Q": ["生物学", "基因", "进化论", "生态学", "微生物", "植物学", "动物学", "遗传学", "分子生物学", "神经科学"],
    "R": ["医学", "中医", "药学", "营养学", "健康", "临床医学", "外科", "内科", "中药", "针灸", "养生", "流行病学"],
    "S": ["农业", "园艺", "林业", "畜牧", "水产", "宠物", "花卉", "盆景"],
    "T": ["编程", "计算机", "Python", "算法", "人工智能", "机器学习", "深度学习", "网络安全", "数据库", "软件工程", "操作系统", "Linux", "大数据", "电子", "机械", "材料", "能源", "自动化", "机器人", "建筑学", "室内设计", "土木工程", "通信"],
    "U": ["铁路", "交通", "汽车", "船舶", "高铁", "地铁", "桥梁", "隧道", "公路"],
    "V": ["航空", "航天", "飞行器", "火箭", "卫星", "NASA", "太空", "登月"],
    "X": ["环境", "环保", "可持续发展", "气候变化", "环境科学", "污染防治", "生态保护"],
    "Z": ["百科全书", "丛书", "年鉴", "辞典", "工具书", "全集"],
}

# ── 当当分类 → CLC 映射 ─────────────────────────────────
DANGDANG_CATEGORY_TO_CLC = {
    "小说": "I", "文学": "I", "青春文学": "I", "动漫": "J",
    "艺术": "J", "摄影": "J", "音乐": "J", "绘画": "J", "书法": "J",
    "传记": "K", "历史": "K", "地理": "K", "旅游": "K", "地图": "P",
    "心理学": "B", "哲学": "B", "宗教": "B", "哲学/宗教": "B",
    "社会科学": "C", "社会学": "C", "管理": "C", "社会科学总论": "C",
    "政治": "D", "政治/军事": "D", "法律": "D", "军事": "E",
    "经济": "F", "金融": "F", "投资": "F", "理财": "F", "管理": "F",
    "文化": "G", "教育": "G", "体育": "G", "新闻传播": "G",
    "语言": "H", "外语": "H", "英语": "H", "日语": "H",
    "自然科学": "N", "科普": "N", "科学": "N",
    "数学": "O", "物理": "O", "化学": "O", "数理化": "O",
    "天文": "P", "地球科学": "P", "地质学": "P", "气象": "P",
    "生物": "Q", "生物科学": "Q", "医学": "R", "医药卫生": "R", "中医": "R",
    "农业": "S", "林业": "S", "畜牧": "S", "水产": "S",
    "计算机": "T", "计算机/网络": "T", "工业技术": "T", "建筑": "T",
    "电子": "T", "通信": "T", "机械": "T", "能源": "T", "自动化": "T",
    "交通": "U", "汽车": "U", "铁路": "U",
    "航空": "V", "航天": "V",
    "环境": "X", "环保": "X", "环境科学": "X",
    "工具书": "Z", "百科全书": "Z", "辞典": "Z",
    "考试": "G", "教材": "G", "中小学": "G", "少儿": "I", "童书": "I",
    "生活": "Z", "美食": "T", "手工": "J", "家居": "T",
    "成功": "B", "励志": "B", "修养": "B",
}

# ── 当当分类ID → 对应CLC类别 (用于URL发现和调度) ────────
DANGDANG_CATEGORIES = [
    # (cat_id, cat_name, target_clc)
    ("01.03.00.00.00.00", "小说", "I"),
    ("01.04.00.00.00.00", "文学", "I"),
    ("01.11.00.00.00.00", "青春文学", "I"),
    ("01.13.00.00.00.00", "动漫/幽默", "J"),
    ("01.09.00.00.00.00", "艺术", "J"),
    ("01.16.00.00.00.00", "传记", "K"),
    ("01.12.00.00.00.00", "历史", "K"),
    ("01.10.00.00.00.00", "旅游/地图", "K"),
    ("01.21.00.00.00.00", "哲学/宗教", "B"),
    ("01.25.00.00.00.00", "心理学", "B"),
    ("01.18.00.00.00.00", "社会科学", "C"),
    ("01.06.00.00.00.00", "管理", "F"),
    ("01.17.00.00.00.00", "政治/军事", "D"),
    ("01.19.00.00.00.00", "法律", "D"),
    ("01.07.00.00.00.00", "经济", "F"),
    ("01.08.00.00.00.00", "投资理财", "F"),
    ("01.20.00.00.00.00", "文化", "G"),
    ("01.02.00.00.00.00", "中小学教辅", "G"),
    ("01.01.00.00.00.00", "童书", "I"),
    ("01.26.00.00.00.00", "外语", "H"),
    ("01.23.00.00.00.00", "科普", "N"),
    ("01.27.00.00.00.00", "自然科学", "N"),
    ("01.32.00.00.00.00", "计算机/网络", "T"),
    ("01.15.00.00.00.00", "建筑", "T"),
    ("01.35.00.00.00.00", "工业技术", "T"),
    ("01.22.00.00.00.00", "医学", "R"),
    ("01.33.00.00.00.00", "农业/林业", "S"),
    ("01.14.00.00.00.00", "家居", "T"),
    ("01.24.00.00.00.00", "体育/运动", "G"),
    ("01.05.00.00.00.00", "考试", "G"),
    ("01.28.00.00.00.00", "工具书", "Z"),
    ("01.29.00.00.00.00", "成功/励志", "B"),
    ("01.30.00.00.00.00", "生活", "Z"),
]

# ── User-Agent 池 ─────────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:128.0) Gecko/20100101 Firefox/128.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:128.0) Gecko/20100101 Firefox/128.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 Edg/125.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 OPR/111.0.0.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64; rv:128.0) Gecko/20100101 Firefox/128.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.3 Safari/605.1.15",
]


# ╔══════════════════════════════════════════════════════════════╗
# ║                    SourceSession                           ║
# ╚══════════════════════════════════════════════════════════════╝
class SourceSession:
    """通用HTTP会话 — 支持多源独立配置"""

    def __init__(self, name: str, cookie_dir: Path = DATA_DIR, warmup_url: str = ""):
        self.name = name
        self.cookie_file = cookie_dir / f"cookies_{name}.json"
        self.last_request_time = 0.0
        self.consecutive_failures = 0
        self.blocked_until = 0.0
        self.warmup_url = warmup_url
        self._build_client()

    def _build_client(self):
        ua = random.choice(USER_AGENTS)
        cookies = self._load_cookies()
        self.client = httpx.Client(
            headers={
                "User-Agent": ua,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                "Accept-Encoding": "gzip, deflate",
                "Connection": "keep-alive",
            },
            cookies=cookies or {},
            timeout=30.0,
            follow_redirects=True,
            http2=False,
        )

    def _load_cookies(self) -> dict | None:
        if self.cookie_file.exists():
            try:
                mtime = os.path.getmtime(self.cookie_file)
                if time.time() - mtime > 3600:
                    self.cookie_file.unlink()
                    return None
                with open(self.cookie_file, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return None
        return None

    def _save_cookies(self):
        try:
            cookies = dict(self.client.cookies)
            if cookies:
                with open(self.cookie_file, "w", encoding="utf-8") as f:
                    json.dump(cookies, f, ensure_ascii=False)
        except Exception:
            pass

    def _enforce_delay(self):
        elapsed = time.time() - self.last_request_time
        cooldown_remaining = self.blocked_until - time.time()
        if cooldown_remaining > 0:
            logger.info(f"[{self.name}] 封禁冷却中, 剩余 {cooldown_remaining:.0f}s ...")
            time.sleep(cooldown_remaining)
            self.blocked_until = 0.0

        required = REQUEST_DELAY_MIN + random.random() * (REQUEST_DELAY_MAX - REQUEST_DELAY_MIN)
        if self.consecutive_failures > 0:
            required *= (1 + 0.5 * self.consecutive_failures)
        if elapsed < required:
            time.sleep(required - elapsed)

    def _rotate_if_needed(self):
        if self.consecutive_failures >= 3:
            cooldown = random.uniform(60, 180)
            logger.info(f"[{self.name}] 连续失败{self.consecutive_failures}次 → 更换UA, 冷却{cooldown:.0f}s")
            self.blocked_until = time.time() + cooldown
            self._build_client()
            self.consecutive_failures = 0

    def is_blocked(self) -> bool:
        return self.blocked_until > time.time()

    def get(self, url: str, referer: str = "", encoding: str = "") -> httpx.Response | None:
        """发送GET请求（含重试+反爬处理）"""
        self._rotate_if_needed()
        headers = {}
        if referer:
            headers["Referer"] = referer

        for attempt in range(MAX_RETRIES):
            self._enforce_delay()
            try:
                resp = self.client.get(url, headers=headers or None)
                self.last_request_time = time.time()

                if encoding:
                    resp.encoding = encoding
                elif resp.encoding in ("", "ISO-8859-1", "latin-1"):
                    # 自动检测中文编码
                    ct = resp.headers.get("content-type", "")
                    m = re.search(r"charset=([\w-]+)", ct)
                    if m:
                        resp.encoding = m.group(1)
                    else:
                        # 尝试从内容检测
                        sample = resp.content[:2000]
                        m2 = re.search(rb'charset[="\s]+([\w-]+)', sample, re.I)
                        if m2:
                            resp.encoding = m2.group(1).decode("ascii", errors="ignore")

                if resp.status_code == 200:
                    self.consecutive_failures = max(0, self.consecutive_failures - 1)
                    self._save_cookies()
                    return resp
                elif resp.status_code == 404:
                    return None
                elif resp.status_code in (418, 429, 403):
                    if resp.status_code == 418:
                        delay = RETRY_BASE_DELAY * (3 ** attempt) + random.uniform(5, 15)
                        logger.warning(f"[{self.name}] HTTP 418 → 等待 {delay:.0f}s (第{attempt+1}次)")
                    else:
                        delay = RETRY_BASE_DELAY * (2 ** attempt) + random.uniform(0, 2)
                        logger.warning(f"[{self.name}] HTTP {resp.status_code} → 等待 {delay:.0f}s (第{attempt+1}次)")
                    time.sleep(delay)
                    self.consecutive_failures += 1
                    self._build_client()
                    if attempt >= 2:
                        return None
                elif resp.status_code >= 500:
                    delay = RETRY_BASE_DELAY * (2 ** attempt)
                    logger.warning(f"[{self.name}] HTTP {resp.status_code} → 等待 {delay:.0f}s 后重试")
                    time.sleep(delay)
                else:
                    logger.warning(f"[{self.name}] HTTP {resp.status_code} → 跳过: {url[:80]}")
                    return None
            except (httpx.TimeoutException, httpx.ConnectError, httpx.RemoteProtocolError) as e:
                if attempt >= 2:
                    logger.warning(f"[{self.name}] 网络错误连续3次, 跳过: {url[:80]}")
                    return None
                delay = 2.0 * (2 ** attempt)
                logger.warning(f"[{self.name}] 网络错误: {e} → 等待 {delay:.0f}s (第{attempt+1}次)")
                time.sleep(delay)

        logger.error(f"[{self.name}] 重试{MAX_RETRIES}次后仍失败: {url[:80]}")
        return None


# ╔══════════════════════════════════════════════════════════════╗
# ║                   DoubanParser                             ║
# ╚══════════════════════════════════════════════════════════════╝
class DoubanParser:
    """豆瓣页面HTML解析器"""

    @staticmethod
    def extract_douban_id(url: str) -> str | None:
        m = re.search(r"/subject/(\d+)/?", url)
        return m.group(1) if m else None

    @classmethod
    def parse_listing_page(cls, html: str) -> list[dict]:
        soup = BeautifulSoup(html, "lxml")
        books = []
        for item in soup.select("li.subject-item"):
            link = item.select_one("div.info h2 a")
            if not link:
                continue
            title = link.get_text(strip=True)
            href = link.get("href", "")
            douban_id = cls.extract_douban_id(href)
            if not douban_id:
                continue
            pub_info = item.select_one("div.pub")
            pub_text = pub_info.get_text(strip=True) if pub_info else ""
            rating_el = item.select_one("span.rating_nums")
            rating = rating_el.get_text(strip=True) if rating_el else ""
            books.append({
                "douban_id": douban_id,
                "title": title,
                "pub_info": pub_text,
                "rating": rating,
                "url": href if href.startswith("http") else DOUBAN_BASE + href,
            })
        return books

    @classmethod
    def parse_book_page(cls, html: str, url: str) -> dict:
        soup = BeautifulSoup(html, "lxml")
        douban_id = cls.extract_douban_id(url) or ""

        # 书名
        title = ""
        for sel in ['#wrapper h1 span[property="v:itemreviewed"]', "#wrapper h1", '[property="v:itemreviewed"]']:
            el = soup.select_one(sel)
            if el and el.get_text(strip=True):
                title = el.get_text(strip=True)
                break

        # 简介
        description = ""
        for sel in ["#link-report .intro", "#link-report .intro p",
                     ".related_info .indent .intro", '.related_info .indent span[class*="intro"]', "#link-report"]:
            els = soup.select(sel)
            if els:
                parts = [el.get_text(strip=True) for el in els if el.get_text(strip=True) and len(el.get_text(strip=True)) > 5]
                if parts:
                    description = "\n".join(parts)
                    break
        if not description:
            intro_div = soup.select_one("#link-report")
            if intro_div:
                for hidden in intro_div.select(".all.hidden"):
                    hidden.extract()
                desc = intro_div.get_text(strip=True)
                if desc and len(desc) > 5:
                    description = desc

        # 标签
        tags = []
        tag_section = soup.select_one("#db-tags-section") or soup.select_one(".tags")
        if tag_section:
            for a in tag_section.select("a.tag"):
                tag_text = a.get_text(strip=True)
                if tag_text:
                    tags.append(tag_text)

        # 作者/出版社/出版年/ISBN
        author = publisher = pub_year = isbn = ""
        info_div = soup.select_one("#info")
        if info_div:
            info_text = info_div.get_text("|", strip=True)
            author_el = info_div.select_one('a[href*="author"]') or info_div.select_one('span a[href*="search"]')
            if author_el:
                author = author_el.get_text(strip=True)
            else:
                m = re.search(r"作者[：:]\s*(.+?)(?:\||$)", info_text)
                if not m:
                    m = re.search(r"作者[：:]\s*(.+?)$", info_text)
                if m:
                    author = m.group(1).strip().split("|")[0].strip()
            m = re.search(r"出版社[：:]\s*(.+?)(?:\||$)", info_text)
            if m:
                publisher = m.group(1).strip()
            m = re.search(r"出版年[：:]\s*(.+?)(?:\||$)", info_text)
            if m:
                pub_year = m.group(1).strip()
            m = re.search(r"ISBN[：:]\s*([\d\-Xx]+)", info_text)
            if m:
                isbn = m.group(1).strip()

        # 评分
        rating = ""
        rating_el = soup.select_one("strong.ll.rating_num") or soup.select_one('[property="v:average"]')
        if rating_el:
            rating = rating_el.get_text(strip=True)

        rating_count = ""
        rc_el = soup.select_one("a.rating_people span") or soup.select_one('[property="v:votes"]')
        if rc_el:
            rating_count = rc_el.get_text(strip=True)

        return {
            "douban_id": douban_id, "title": title, "author": author,
            "publisher": publisher, "pub_year": pub_year, "isbn": isbn,
            "description": description, "douban_tags": tags,
            "rating": rating, "rating_count": rating_count,
            "source_url": url, "source_name": "douban",
            "scraped_at": datetime.now().isoformat(),
        }


# ╔══════════════════════════════════════════════════════════════╗
# ║                   DangdangParser                           ║
# ╚══════════════════════════════════════════════════════════════╝
class DangdangParser:
    """当当网HTML解析器 — 编码GBK"""

    @staticmethod
    def _clean_html(html_bytes: bytes) -> str:
        """自动检测编码并解码"""
        sample = html_bytes[:2000]
        m = re.search(rb'charset[="\s]+([\w-]+)', sample, re.I)
        encoding = m.group(1).decode("ascii", errors="ignore") if m else "gbk"
        return html_bytes.decode(encoding, errors="ignore")

    @classmethod
    def parse_listing_page(cls, html_bytes: bytes) -> list[dict]:
        """解析分类列表页"""
        html = cls._clean_html(html_bytes)
        soup = BeautifulSoup(html, "lxml")
        books = []
        for item in soup.select('li[id^="p"]'):
            product_id = item.get("id", "").lstrip("p")
            if not product_id:
                continue
            name_el = item.select_one('p.name a')
            title = name_el.get("title", "") or name_el.get_text(strip=True) if name_el else ""
            if not title:
                continue
            url = f"{DANGDANG_PRODUCT}/{product_id}.html"
            # 列表页的简介
            desc_el = item.select_one('p.detail')
            desc = desc_el.get_text(strip=True) if desc_el else ""
            # 作者和出版社
            author_els = item.select('p.search_book_author span a')
            author = author_els[0].get_text(strip=True) if author_els else ""
            # 评分
            star_span = item.select_one('span.search_star_black span')
            rating = ""
            if star_span:
                w = star_span.get("style", "")
                m = re.search(r'(\d+)', w)
                if m:
                    rating = str(int(m.group(1)) / 20)
            books.append({
                "product_id": product_id,
                "title": title.strip(),
                "author": author,
                "description_hint": desc,
                "url": url,
                "rating": rating,
            })
        return books

    @classmethod
    def parse_book_page(cls, html_bytes: bytes, url: str) -> dict:
        """解析书籍详情页"""
        html = cls._clean_html(html_bytes)
        soup = BeautifulSoup(html, "lxml")
        product_id = ""
        for pat in [r'/product/(\d+)', r'/(\d+)\.html', r'product_id=(\d+)']:
            m = re.search(pat, url)
            if m:
                product_id = m.group(1)
                break

        # 书名
        title_el = soup.select_one('h1[dd_name]') or soup.select_one('div.name_info h1') or soup.select_one('h1')
        title = title_el.get_text(strip=True) if title_el else ""

        # 分类面包屑 — 过滤掉书名(过长)和特殊字符
        bread = soup.select_one('div.breadcrumb')
        bread_text = bread.get_text(strip=True) if bread else ""
        categories = []
        for b in bread_text.split(">"):
            b = b.strip()
            if not b or b == "图书":
                continue
            if len(b) > 15 or "\xa0" in b:
                continue
            categories.append(b)

        # 简介
        description = ""
        desc_area = soup.select_one('div#detail_all') or soup.select_one('span.descrip') or soup.select_one('div.descrip')
        if desc_area:
            desc_all = desc_area.get_text(" ", strip=True)
            if "本商品暂无详情" not in desc_all:
                # 尝试匹配"内容简介"区域
                m = re.search(r'内容简介[：:]?\s*(.+?)(?:作者简介|目录|媒体评论|免费下载|下载免费|当当读书APP|所属分类|开\s*本)', desc_all)
                if m and len(m.group(1).strip()) > 10:
                    description = m.group(1).strip()
                else:
                    parts = re.split(r'(?:[。；]|\n)\s*', desc_all)
                    meaningful = [p.strip() for p in parts
                                  if len(p.strip()) > 20 and '下载' not in p
                                  and '当当' not in p and '扫描' not in p
                                  and '手机' not in p and 'ISBN' not in p
                                  and '开 本' not in p and '纸 张' not in p
                                  and '包 装' not in p and '所属分类' not in p
                                  and 'APP' not in p and '阅读器' not in p]
                    description = meaningful[0] if meaningful else ""

        # ISBN/出版社/分类/出版时间 — 从 ul.key 提取
        isbn = author = publisher = pub_year = ""
        info_ul = soup.select_one('ul.key')
        if info_ul:
            info_text = info_ul.get_text(" ", strip=True)
            m = re.search(r'ISBN[：:]\s*([\d\-Xx]+)', info_text)
            if m:
                isbn = m.group(1).strip()
            # 所属分类 (可能比面包屑更完整)
            m_cat = re.search(r'所属分类[：:]\s*(.+?)(?:下载|当当|$)', info_text)
            if m_cat:
                raw_cat = m_cat.group(1).strip()
                cats = [c.strip() for c in raw_cat.split(">") if c.strip() and c.strip() != "图书" and len(c.strip()) <= 15]
                if cats and len(cats) > len(categories):
                    categories = cats
            # 出版社
            m_pub = re.search(r'出版社[：:]\s*(.+?)(?:\s|$)', info_text)
            if m_pub:
                publisher = m_pub.group(1).strip()
            if not publisher:
                for li in info_ul.select('li'):
                    text = li.get_text(strip=True)
                    if "出版社" in text:
                        publisher = text.replace("出版社：", "").replace("出版社:", "").strip()
                        break
            # 出版时间
            m2 = re.search(r'出版时间[：:]\s*(\d{4})', info_text)
            if m2:
                pub_year = m2.group(1)

        # 作者: 优先用单品作者标签
        if not author:
            author_el = soup.select_one('a[dd_name="单品作者"]') or soup.select_one('span.t1 a')
            if author_el:
                author = author_el.get_text(strip=True)
        if not author and info_ul:
            for li in info_ul.select('li'):
                text = li.get_text(strip=True)
                if "作者" in text:
                    author = text.replace("作者：", "").replace("作者:", "").strip()
                    break

        # 评分
        rating = ""
        star_el = soup.select_one('div.star') or soup.select_one('span.level')
        if star_el:
            style = star_el.get("style", "")
            m = re.search(r'(\d+)', style)
            if m:
                rating = str(int(m.group(1)) / 20)

        return {
            "douban_id": f"dd_{product_id}",  # 统一用douban_id字段名存储
            "product_id": product_id,
            "title": title,
            "author": author,
            "publisher": publisher,
            "pub_year": pub_year,
            "isbn": isbn,
            "description": description,
            "douban_tags": categories,  # 复用字段: 当当分类面包屑
            "rating": rating,
            "rating_count": "",
            "source_url": url,
            "source_name": "dangdang",
            "scraped_at": datetime.now().isoformat(),
        }

    @staticmethod
    def map_categories_to_clc(categories: list[str]) -> list[dict]:
        """当当分类面包屑 → CLC分类号 (长键优先，更精确)"""
        results = []
        # 按key长度降序排列，长匹配优先
        sorted_keys = sorted(DANGDANG_CATEGORY_TO_CLC.keys(), key=len, reverse=True)
        for cat in categories:
            for key in sorted_keys:
                if key in cat:
                    clc = DANGDANG_CATEGORY_TO_CLC[key]
                    # 避免在结果列表中出现完全重复的(clc, sub)
                    if not any(r["clc_code"] == clc and r["sub"] == cat for r in results):
                        results.append({"clc_code": clc, "sub": cat, "confidence": 0.75})
                    break  # 只取最长匹配
        if not results:
            results.append({"clc_code": "Z", "sub": "综合性图书", "confidence": 0.0})
        return results


# ╔══════════════════════════════════════════════════════════════╗
# ║                       CLCMapper                            ║
# ╚══════════════════════════════════════════════════════════════╝
class CLCMapper:
    """标签 → 中图法分类号映射器（共用）"""

    def __init__(self, mapping_path: Path = TAG_MAP_FILE):
        with open(mapping_path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        self.tag_map = {k: v for k, v in raw.items() if not k.startswith("_")}
        self._known_tags = list(self.tag_map.keys())

    def map_tags_to_clc(self, tags: list[str]) -> list[dict]:
        if not tags:
            return [{"clc_code": "Z", "sub": "综合性图书", "confidence": 0.0}]
        scores: dict[str, float] = {}
        sub_map: dict[str, str] = {}
        for tag in tags:
            tag_clean = tag.strip()
            if not tag_clean:
                continue
            mappings = self._lookup(tag_clean)
            for m in mappings:
                key = m["clc"]
                scores[key] = scores.get(key, 0.0) + m["confidence"]
                if key not in sub_map:
                    sub_map[key] = m.get("sub", "")
        if not scores:
            return [{"clc_code": "Z", "sub": "综合性图书", "confidence": 0.0}]
        total = sum(scores.values()) or 1.0
        ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)
        return [
            {"clc_code": code, "sub": sub_map.get(code, ""), "confidence": round(score / total, 4)}
            for code, score in ranked
        ]

    def _lookup(self, tag: str) -> list[dict]:
        if tag in self.tag_map:
            return self.tag_map[tag]
        best_ratio = 0.0
        best_entries = []
        for known, entries in self.tag_map.items():
            ratio = SequenceMatcher(None, tag, known).ratio()
            if ratio >= 0.80 and ratio > best_ratio:
                best_ratio = ratio
                best_entries = [
                    {"clc": e["clc"], "sub": e.get("sub", ""), "confidence": e["confidence"] * ratio}
                    for e in entries
                ]
        return best_entries


# ╔══════════════════════════════════════════════════════════════╗
# ║                   CheckpointManager                        ║
# ╚══════════════════════════════════════════════════════════════╝
class CheckpointManager:
    """断点续传状态管理"""

    @staticmethod
    def save(total: int, category_counts: dict, tag_cursor: dict = None,
             dd_cursor: dict = None, source_states: dict = None):
        data = {
            "total_scraped": total,
            "category_counts": dict(category_counts),
            "tag_cursor": tag_cursor or {},
            "dd_cursor": dd_cursor or {},
            "source_states": source_states or {},
            "last_updated": datetime.now().isoformat(),
        }
        tmp = CHECKPOINT_FILE.with_suffix(".tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        tmp.replace(CHECKPOINT_FILE)

    @staticmethod
    def load() -> dict:
        if not CHECKPOINT_FILE.exists():
            return {"total_scraped": 0, "category_counts": {}, "tag_cursor": {},
                    "dd_cursor": {}, "source_states": {}}
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        for key in ("tag_cursor", "dd_cursor", "source_states"):
            if key not in data:
                data[key] = {} if key != "source_states" else {}
        return data


# ╔══════════════════════════════════════════════════════════════╗
# ║                     DedupFilter                            ║
# ╚══════════════════════════════════════════════════════════════╝
class DedupFilter:
    """SQLite持久化去重"""

    def __init__(self, db_path: Path = SEEN_DB):
        self.conn = sqlite3.connect(str(db_path))
        self.conn.execute("CREATE TABLE IF NOT EXISTS seen (douban_id TEXT PRIMARY KEY, scraped_at TEXT)")
        self.conn.execute("CREATE INDEX IF NOT EXISTS idx_seen_id ON seen(douban_id)")
        self.conn.commit()
        self._count = 0

    def is_duplicate(self, douban_id: str) -> bool:
        row = self.conn.execute("SELECT 1 FROM seen WHERE douban_id = ?", (douban_id,)).fetchone()
        return row is not None

    def add(self, douban_id: str):
        self.conn.execute("INSERT OR IGNORE INTO seen VALUES (?, ?)", (douban_id, datetime.now().isoformat()))
        self._count += 1
        if self._count % 100 == 0:
            self.conn.commit()

    def count(self) -> int:
        row = self.conn.execute("SELECT COUNT(*) FROM seen").fetchone()
        return row[0] if row else 0

    def close(self):
        self.conn.commit()
        self.conn.close()


# ╔══════════════════════════════════════════════════════════════╗
# ║                     ExcelWriter                            ║
# ╚══════════════════════════════════════════════════════════════╝
class ExcelWriter:
    """缓冲写入Excel (.xlsx) — 支持断点续写"""

    COLUMNS = [
        "douban_id", "书名", "作者", "出版社", "出版年", "ISBN",
        "内容简介", "豆瓣标签", "中图法分类号", "次要分类号",
        "分类置信度", "豆瓣评分", "来源URL", "数据来源", "抓取时间",
    ]

    def __init__(self, filepath: Path, resume: bool = False):
        self.filepath = filepath
        self._buffer: list[list] = []

        if resume and filepath.exists():
            # 续写模式：加载已有文件，不覆盖
            from openpyxl import load_workbook
            self.wb = load_workbook(str(filepath))
            self.ws = self.wb.active
            self._total_written = self.ws.max_row - 1  # 减去表头
            # 处理列不兼容：如果旧文件列数与新列数不同，在末尾补齐head
            existing_cols = self.ws.max_column
            if existing_cols < len(self.COLUMNS):
                for ci, col_name in enumerate(self.COLUMNS[existing_cols:], start=existing_cols):
                    self.ws.cell(row=1, column=ci + 1, value=col_name)
            logger.info(f"续写Excel: 已有 {self._total_written} 条数据")
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "图书数据"
            self.ws.append(self.COLUMNS)
            self._total_written = 0

    def write(self, item: dict):
        row = [
            item.get("douban_id", ""),
            item.get("title", ""),
            item.get("author", ""),
            item.get("publisher", ""),
            item.get("pub_year", ""),
            item.get("isbn", ""),
            item.get("description", ""),
            ";".join(item.get("douban_tags", [])),
            item.get("clc_code", ""),
            ";".join(item.get("clc_secondary", [])),
            item.get("clc_confidence", 0.0),
            item.get("rating", ""),
            item.get("source_url", ""),
            item.get("source_name", ""),
            item.get("scraped_at", ""),
        ]
        self._buffer.append(row)
        if len(self._buffer) >= BATCH_SIZE:
            self._flush()

    def _flush(self):
        for row in self._buffer:
            self.ws.append(row)
        self._total_written += len(self._buffer)
        self._buffer.clear()
        self.wb.save(str(self.filepath))
        logger.info(f"已写入 {self._total_written} 条数据到 {self.filepath.name}")

    def close(self):
        self._flush()
        for col_cells in self.ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            self.ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)
        self.ws.auto_filter.ref = self.ws.dimensions
        self.wb.save(str(self.filepath))
        logger.info(f"文件已保存: {self.filepath} (共 {self._total_written} 条)")


# ╔══════════════════════════════════════════════════════════════╗
# ║                   AdaptiveBalancer                         ║
# ╚══════════════════════════════════════════════════════════════╝
class AdaptiveBalancer:
    """自适应分类平衡调度器"""

    def __init__(self, targets: dict):
        self.targets = targets
        self.current = Counter()

    def update(self, clc_code: str):
        main_cat = clc_code[0] if clc_code else "Z"
        if main_cat in self.targets:
            self.current[main_cat] += 1

    def get_deficit_categories(self) -> list[tuple[str, float]]:
        deficits = []
        for cat, target in self.targets.items():
            current = self.current.get(cat, 0)
            if current < target:
                deficits.append((cat, current / target))
        deficits.sort(key=lambda x: x[1])
        return deficits

    def next_tags(self, n: int = 10) -> list[str]:
        """豆瓣源: 返回最需要爬取的标签"""
        deficits = self.get_deficit_categories()
        tags = []
        for cat, ratio in deficits:
            cat_tags = CATEGORY_TAGS.get(cat, [])
            if cat_tags:
                pick = max(1, min(3, int((1 - ratio) * len(cat_tags) + 1)))
                tags.extend(random.sample(cat_tags, min(pick, len(cat_tags))))
            if len(tags) >= n:
                break
        if len(tags) < n:
            all_tags = [t for ts in CATEGORY_TAGS.values() for t in ts]
            remaining = [t for t in all_tags if t not in tags]
            tags.extend(random.sample(remaining, min(n - len(tags), len(remaining))))
        return tags[:n]

    def next_dangdang_categories(self, n: int = 5) -> list[tuple[str, str, str]]:
        """当当源: 返回最需要爬取的分类 (cat_id, cat_name, clc)"""
        deficits = self.get_deficit_categories()
        deficit_clcs = {cat for cat, _ in deficits}
        # 按缺口筛选
        result = []
        for cat_id, cat_name, clc in DANGDANG_CATEGORIES:
            if clc in deficit_clcs:
                result.append((cat_id, cat_name, clc))
        # 缺口类不够时补充
        remaining = [(c[0], c[1], c[2]) for c in DANGDANG_CATEGORIES if c not in result]
        result.extend(remaining)
        return result[:n]

    def summary(self) -> str:
        lines = []
        for cat in sorted(self.targets.keys()):
            cur = self.current.get(cat, 0)
            tgt = self.targets[cat]
            pct = cur / tgt * 100 if tgt else 0
            bar = "#" * int(pct / 5) + "." * (20 - int(pct / 5))
            lines.append(f"  {cat}: {bar} {cur}/{tgt} ({pct:.0f}%)")
        return "\n".join(lines)


# ╔══════════════════════════════════════════════════════════════╗
# ║                   SourceManager                            ║
# ╚══════════════════════════════════════════════════════════════╝
class SourceManager:
    """多源轮换管理器"""

    SOURCE_ORDER = ["douban", "dangdang"]

    def __init__(self, source_states: dict = None):
        self.blocked_until: dict[str, float] = {}
        if source_states:
            for name, state in source_states.items():
                if state.get("blocked_until", 0) > time.time():
                    self.blocked_until[name] = state["blocked_until"]

    def active_source(self) -> str:
        """返回第一个可用源名称"""
        for name in self.SOURCE_ORDER:
            until = self.blocked_until.get(name, 0)
            if until <= time.time():
                return name
        # 全部被封，等冷却最快的
        min_name = min(self.blocked_until, key=self.blocked_until.get)
        wait = self.blocked_until[min_name] - time.time()
        if wait > 0:
            logger.info(f"所有源均被封禁, 等待 {wait:.0f}s (最快恢复: {min_name})")
            time.sleep(wait)
        self.blocked_until.pop(min_name, None)
        return min_name

    def mark_blocked(self, name: str, cooldown_sec: float):
        until = time.time() + cooldown_sec
        self.blocked_until[name] = until
        logger.warning(f"源 [{name}] 被封, 冷却 {cooldown_sec:.0f}s (恢复时间: {datetime.fromtimestamp(until).strftime('%H:%M:%S')})")

    def get_state(self) -> dict:
        return {name: {"blocked_until": self.blocked_until.get(name, 0)}
                for name in self.SOURCE_ORDER}


# ╔══════════════════════════════════════════════════════════════╗
# ║                      BookScraper                           ║
# ╚══════════════════════════════════════════════════════════════╝
class BookScraper:
    """主调度器 — 多源版"""

    def __init__(self, max_books: int = MAX_BOOKS_DEFAULT, resume: bool = False):
        self.max_books = max_books
        self.mapper = CLCMapper()
        self.dedup = DedupFilter()
        self.writer = ExcelWriter(OUTPUT_FILE, resume=resume)
        self.balancer = AdaptiveBalancer(CATEGORY_TARGETS)
        self.total = 0

        # 豆瓣源状态
        self.tag_cursor: dict[str, int] = {}
        self.bad_tags: set[str] = set()
        self.douban_session = SourceSession("douban", DATA_DIR)
        self.douban_parser = DoubanParser()

        # 当当源状态
        self.dd_cursor: dict[str, int] = {}  # cat_id → page_index
        self.dangdang_session = SourceSession("dangdang", DATA_DIR)
        self.dangdang_parser = DangdangParser()

        # SourceManager
        self.source_mgr = SourceManager()

        if resume:
            state = CheckpointManager.load()
            self.total = state.get("total_scraped", 0)
            self.balancer.current = Counter(state.get("category_counts", {}))
            self.tag_cursor = state.get("tag_cursor", {})
            self.dd_cursor = state.get("dd_cursor", {})
            self.source_mgr = SourceManager(state.get("source_states", {}))
            logger.info(f"从断点恢复: 已爬取 {self.total} 条")

    def run(self):
        logger.info(f"目标: {self.max_books} 条 | 当前: {self.total} 条 | 源: douban, dangdang")
        try:
            while self.total < self.max_books:
                source_name = self.source_mgr.active_source()
                if source_name == "douban":
                    self._run_douban()
                elif source_name == "dangdang":
                    self._run_dangdang()
        except KeyboardInterrupt:
            logger.info("\n收到中断信号，保存进度...")
            self._save_checkpoint()
        finally:
            self.writer.close()
            self.dedup.close()
            self._print_summary()

    def _run_douban(self):
        """豆瓣源爬取循环"""
        session = self.douban_session
        parser = self.douban_parser
        while self.total < self.max_books and not session.is_blocked():
            tags = [t for t in self.balancer.next_tags(n=15) if t not in self.bad_tags]
            if not tags:
                break
            for tag in tags:
                if self.total >= self.max_books or session.is_blocked():
                    break
                self._crawl_douban_tag(session, parser, tag)
            # 标签不够时随机ID补充
            if self.total < self.max_books and not session.is_blocked():
                self._crawl_douban_random(session, parser, min(5000, self.max_books - self.total))

        if session.is_blocked():
            self.source_mgr.mark_blocked("douban", random.uniform(600, 1800))

    def _crawl_douban_tag(self, session, parser, tag: str):
        start = self.tag_cursor.get(tag, 0)
        base_url = f"{DOUBAN_TAG}/{tag}"
        while self.total < self.max_books and not session.is_blocked():
            url = f"{base_url}?start={start}&type=T"
            logger.info(f"[douban][{self.total}/{self.max_books}] 标签: {tag} (start={start})")
            resp = session.get(url, referer=DOUBAN_BASE + "/")
            if not resp:
                if start == 0:
                    self.bad_tags.add(tag)
                break
            if "/misc/sorry" in str(resp.url):
                logger.warning(f"豆瓣限流 → 标签 {tag} 暂停")
                session.blocked_until = time.time() + random.uniform(120, 300)
                self.bad_tags.add(tag)
                break
            books = parser.parse_listing_page(resp.text)
            if not books:
                if start == 0:
                    self.bad_tags.add(tag)
                break
            for book_summary in books:
                if self.total >= self.max_books or session.is_blocked():
                    break
                if self.dedup.is_duplicate(book_summary["douban_id"]):
                    continue
                self._scrape_douban_book(session, parser, book_summary, tag)
            start += 20
            self.tag_cursor[tag] = start
            if start > 980:
                break

    def _scrape_douban_book(self, session, parser, book_summary: dict, context: str):
        """爬取单本豆瓣书"""
        url, douban_id = book_summary["url"], book_summary["douban_id"]
        resp = session.get(url, referer=DOUBAN_BASE + "/tag/" if context not in ("random", "") else DOUBAN_BASE + "/")
        if not resp:
            return
        try:
            data = parser.parse_book_page(resp.text, url)
        except Exception as e:
            logger.error(f"解析失败 {url}: {e}")
            return
        if not data.get("title"):
            return

        tags = data.get("douban_tags", [])
        if not tags and context and context not in ("random", ""):
            tags = [context]
        mappings = self.mapper.map_tags_to_clc(tags)
        data["clc_code"] = mappings[0]["clc_code"] if mappings else "Z"
        data["clc_secondary"] = [m["clc_code"] for m in mappings[1:4]] if len(mappings) > 1 else []
        data["clc_confidence"] = mappings[0]["confidence"] if mappings else 0.0

        self._commit_book(data, douban_id)

    def _crawl_douban_random(self, session, parser, count: int):
        """豆瓣随机ID扫描"""
        attempted = 0
        while self.total < self.max_books and attempted < count * 3 and not session.is_blocked():
            douban_id = str(random.randint(1_000_000, 37_000_000))
            if self.dedup.is_duplicate(douban_id):
                attempted += 1
                continue
            url = f"{DOUBAN_SUBJECT}/{douban_id}/"
            self._scrape_douban_book(session, parser, {"url": url, "douban_id": douban_id}, "random")
            attempted += 1
            if attempted % 100 == 0:
                logger.info(f"  随机扫描: 尝试 {attempted}, 成功收录 {self.total}")

    def _run_dangdang(self):
        """当当源爬取循环"""
        session = self.dangdang_session
        parser = self.dangdang_parser
        while self.total < self.max_books and not session.is_blocked():
            categories = self.balancer.next_dangdang_categories(n=5)
            if not categories:
                break
            for cat_id, cat_name, clc_target in categories:
                if self.total >= self.max_books or session.is_blocked():
                    break
                self._crawl_dangdang_category(session, parser, cat_id, cat_name)
        if session.is_blocked():
            self.source_mgr.mark_blocked("dangdang", random.uniform(300, 900))

    def _crawl_dangdang_category(self, session, parser, cat_id: str, cat_name: str):
        """爬取当当一个分类下的书籍"""
        page = self.dd_cursor.get(cat_id, 1)
        while self.total < self.max_books and not session.is_blocked() and page <= 100:
            url = f"{DANGDANG_BASE}/cp{cat_id}.html?page_index={page}"
            logger.info(f"[dangdang][{self.total}/{self.max_books}] 分类: {cat_name} (第{page}页)")
            resp = session.get(url, referer="http://book.dangdang.com/")
            if not resp:
                if page == 1:
                    logger.warning(f"当当分类 {cat_name} ({cat_id}) 无响应, 跳过")
                break
            books = parser.parse_listing_page(resp.content)
            if not books:
                break
            for book_summary in books:
                if self.total >= self.max_books or session.is_blocked():
                    break
                pid = book_summary["product_id"]
                dd_id = f"dd_{pid}"
                if self.dedup.is_duplicate(dd_id):
                    continue
                # 列表页的描述作为回退
                hint = book_summary.get("description_hint", "")
                self._scrape_dangdang_book(session, parser, book_summary, cat_name, hint)
            page += 1
            self.dd_cursor[cat_id] = page

    def _scrape_dangdang_book(self, session, parser, book_summary: dict, category: str, desc_hint: str):
        """爬取单本当当书"""
        url, pid = book_summary["url"], book_summary["product_id"]
        dd_id = f"dd_{pid}"
        resp = session.get(url, referer=DANGDANG_BASE + "/")
        if not resp:
            return
        try:
            data = parser.parse_book_page(resp.content, url)
        except Exception as e:
            logger.error(f"当当解析失败 {url}: {e}")
            return
        if not data.get("title"):
            return

        # 如果详情页没有简介，用列表页的
        if not data.get("description") and desc_hint:
            data["description"] = desc_hint

        # CLC分类
        categories = data.get("douban_tags", [])
        if not categories and category:
            categories = [category]
        mappings = DangdangParser.map_categories_to_clc(categories)
        # 如果直接映射失败，尝试用CLCMapper模糊匹配
        if not mappings or mappings[0]["confidence"] < 0.3:
            tag_mappings = self.mapper.map_tags_to_clc(categories)
            if tag_mappings and tag_mappings[0]["confidence"] > 0:
                mappings = tag_mappings
        data["clc_code"] = mappings[0]["clc_code"] if mappings else "Z"
        data["clc_secondary"] = [m["clc_code"] for m in mappings[1:4]] if len(mappings) > 1 else []
        data["clc_confidence"] = mappings[0]["confidence"] if mappings else 0.0

        self._commit_book(data, dd_id)

    def _commit_book(self, data: dict, book_id: str):
        """统一写入"""
        self.writer.write(data)
        self.dedup.add(book_id)
        self.balancer.update(data.get("clc_code", "Z"))
        self.total += 1

        if self.total % 100 == 0:
            logger.info(f"  进度: {self.total}/{self.max_books} | [{data.get('source_name','')}] "
                        f"{data.get('title','')[:20]} → {data.get('clc_code','')} ({data.get('clc_confidence',0):.2f})")

        if self.total % CHECKPOINT_INTERVAL == 0:
            self._save_checkpoint()
            self._log_distribution()

    def _save_checkpoint(self):
        CheckpointManager.save(
            self.total,
            dict(self.balancer.current),
            self.tag_cursor,
            self.dd_cursor,
            self.source_mgr.get_state(),
        )

    def _log_distribution(self):
        logger.info("当前分类分布:\n" + self.balancer.summary())

    def _print_summary(self):
        logger.info("=" * 60)
        logger.info(f"爬取完成! 总计: {self.total} 条")
        logger.info(f"输出文件: {OUTPUT_FILE}")
        logger.info("分类分布:\n" + self.balancer.summary())


# ── 入口 ──────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="中文图书数据采集系统 (多源版)")
    parser.add_argument("--max", type=int, default=MAX_BOOKS_DEFAULT, help="目标采集数量 (默认100000)")
    parser.add_argument("--resume", action="store_true", help="从断点恢复")
    args = parser.parse_args()

    logger.info("=" * 60)
    logger.info("中文图书数据采集系统 (多源版) 启动")
    logger.info(f"数据源: 豆瓣读书 + 当当网")
    logger.info(f"目标: {args.max} 条 | 输出: {OUTPUT_FILE}")
    logger.info("=" * 60)

    scraper = BookScraper(max_books=args.max, resume=args.resume)
    scraper.run()


if __name__ == "__main__":
    main()
